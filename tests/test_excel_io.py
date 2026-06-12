"""Strict TDD tests for the M5 Excel I/O layer."""

from __future__ import annotations

import os
from pathlib import Path
import re
import shutil

import pytest
from openpyxl import load_workbook

import ema.config as config
import ema.excel_io as excel_io_module
from ema.errors import (
    EmaError,
    UndoError,
    ValidationError as EmaValidationError,
    WorkbookFileMissingError,
    WriteVerificationError,
)
from ema.excel_io import check_writable
from ema.models import AppendResult, ColumnDef, ColumnType, UndoResult, WorkbookSchema


def _assert_file_bytes_match(original_path: Path, copy_path: Path) -> None:
    """Assert that two files exist and contain exactly the same bytes."""

    assert original_path.exists()
    assert copy_path.exists()
    assert copy_path.read_bytes() == original_path.read_bytes()


def _sheet_values(path: Path, sheet_name: str) -> list[list[object | None]]:
    """Return worksheet values as a list of row lists."""

    workbook = load_workbook(path)
    try:
        worksheet = workbook[sheet_name]
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _write_row_using_schema(
    path: Path,
    schema: WorkbookSchema,
    row_index: int,
    values: dict[str, object],
) -> None:
    """Write one row directly into a workbook using schema column positions."""

    workbook = load_workbook(path)
    try:
        worksheet = workbook[schema.sheet]
        columns_by_name = {column.name: column for column in schema.columns}
        for column_name, value in values.items():
            column = columns_by_name[column_name]
            worksheet.cell(row=row_index, column=column.index + 1, value=value)
        workbook.save(path)
    finally:
        workbook.close()


def _sibling_file_names(path: Path) -> set[str]:
    """Return the names of files in the target workbook's parent directory."""

    return {child.name for child in path.parent.iterdir() if child.is_file()}


def _plain_schema() -> WorkbookSchema:
    """Build the schema used for plain workbook append tests."""

    return WorkbookSchema(
        schema_version=config.SCHEMA_VERSION,
        sheet="FoodLog",
        header_row=1,
        table_name=None,
        columns=[
            ColumnDef(name="Date", index=0, type=ColumnType.DATE, required=True),
            ColumnDef(name="Meal", index=1, type=ColumnType.TEXT, required=True),
            ColumnDef(name="Food", index=2, type=ColumnType.TEXT, required=True),
            ColumnDef(name="Calories", index=3, type=ColumnType.NUMBER, required=True),
        ],
        data_start_row=2,
    )


def _formula_schema() -> WorkbookSchema:
    """Build the schema used for formula workbook append tests."""

    return WorkbookSchema(
        schema_version=config.SCHEMA_VERSION,
        sheet="OrderLines",
        header_row=1,
        table_name=None,
        columns=[
            ColumnDef(name="Item", index=0, type=ColumnType.TEXT, required=True),
            ColumnDef(name="Quantity", index=1, type=ColumnType.NUMBER, required=True),
            ColumnDef(name="Price", index=2, type=ColumnType.NUMBER, required=True),
            ColumnDef(
                name="Total",
                index=3,
                type=ColumnType.NUMBER,
                required=False,
                is_formula=True,
            ),
        ],
        data_start_row=2,
    )


def _offset_schema() -> WorkbookSchema:
    """Build the schema used for offset-header workbook append tests."""

    return WorkbookSchema(
        schema_version=config.SCHEMA_VERSION,
        sheet="Bloodwork",
        header_row=3,
        table_name=None,
        columns=[
            ColumnDef(name="Date", index=0, type=ColumnType.DATE, required=True),
            ColumnDef(name="Marker", index=1, type=ColumnType.TEXT, required=True),
            ColumnDef(name="Value", index=2, type=ColumnType.NUMBER, required=True),
            ColumnDef(name="Unit", index=3, type=ColumnType.TEXT, required=True),
        ],
        data_start_row=4,
    )


def _table_schema() -> WorkbookSchema:
    """Build the schema used for Excel Table append tests."""

    return WorkbookSchema(
        schema_version=config.SCHEMA_VERSION,
        sheet="Expenses",
        header_row=1,
        table_name="ExpensesTable",
        columns=[
            ColumnDef(name="Date", index=0, type=ColumnType.DATE, required=True),
            ColumnDef(name="Store", index=1, type=ColumnType.TEXT, required=True),
            ColumnDef(name="Amount", index=2, type=ColumnType.NUMBER, required=True),
            ColumnDef(name="Category", index=3, type=ColumnType.TEXT, required=True),
        ],
        data_start_row=2,
    )


def test_check_writable_passes_for_existing_valid_workbook(plain_wb: Path) -> None:
    """A normal existing workbook should be considered writable."""

    check_writable(plain_wb)


def test_check_writable_raises_for_missing_workbook(tmp_path: Path) -> None:
    """A missing workbook path should raise the typed missing-file exception."""

    missing_path = tmp_path / "missing.xlsx"

    with pytest.raises(WorkbookFileMissingError):
        check_writable(missing_path)


def test_check_writable_raises_for_corrupt_workbook(tmp_path: Path) -> None:
    """A file that is not a real Excel workbook should raise an EMA error."""

    corrupt_path = tmp_path / "corrupt.xlsx"
    corrupt_path.write_bytes(b"this is not a valid xlsx file")

    with pytest.raises(EmaError):
        check_writable(corrupt_path)


def test_check_writable_raises_for_directory_path(tmp_path: Path) -> None:
    """A directory path should not be treated as a writable workbook."""

    directory_path = tmp_path / "not-a-workbook"
    directory_path.mkdir()

    with pytest.raises(WorkbookFileMissingError):
        check_writable(directory_path)


def test_check_writable_passes_for_formula_workbook(formula_wb: Path) -> None:
    """A valid workbook containing formulas should still pass writability checks."""

    check_writable(formula_wb)


def test_check_writable_passes_for_table_workbook(table_wb: Path) -> None:
    """A valid workbook containing an Excel Table should pass writability checks."""

    check_writable(table_wb)


def test_check_writable_passes_for_formatted_workbook(formatted_wb: Path) -> None:
    """A styled workbook should pass writability checks when it opens normally."""

    check_writable(formatted_wb)


def test_backup_creates_backup_file_successfully(
    ema_home: Path, plain_wb: Path
) -> None:
    """A successful backup should create a new workbook copy on disk."""

    backup_path = excel_io_module.backup(plain_wb)

    _assert_file_bytes_match(plain_wb, backup_path)


def test_backup_file_contents_match_original_workbook(
    ema_home: Path, formula_wb: Path
) -> None:
    """Backup contents should exactly match the original workbook bytes."""

    original_bytes = formula_wb.read_bytes()

    backup_path = excel_io_module.backup(formula_wb)

    assert backup_path.exists()
    assert backup_path.read_bytes() == original_bytes


def test_backup_filename_contains_timestamp(ema_home: Path, plain_wb: Path) -> None:
    """Backup filenames should include a timestamp component."""

    backup_path = excel_io_module.backup(plain_wb)

    assert backup_path.exists()
    assert re.match(r"^plain-\d{8}T\d{6}(?:\d+)?Z?\.xlsx$", backup_path.name)
    assert backup_path.read_bytes() == plain_wb.read_bytes()


def test_backup_is_created_inside_ema_backup_directory(
    ema_home: Path, table_wb: Path
) -> None:
    """Backups should be written beneath the configured EMA backup directory."""

    backup_path = excel_io_module.backup(table_wb)

    assert backup_path.exists()
    assert config.BACKUP_DIR in backup_path.parents
    assert backup_path.read_bytes() == table_wb.read_bytes()


def test_multiple_backups_of_same_workbook_create_unique_files(
    ema_home: Path, plain_wb: Path
) -> None:
    """Two backups of the same workbook should produce distinct files."""

    first_backup = excel_io_module.backup(plain_wb)
    second_backup = excel_io_module.backup(plain_wb)

    assert first_backup.exists()
    assert second_backup.exists()
    assert first_backup != second_backup
    assert first_backup.read_bytes() == plain_wb.read_bytes()
    assert second_backup.read_bytes() == plain_wb.read_bytes()


def test_backup_raises_for_missing_workbook(ema_home: Path, tmp_path: Path) -> None:
    """Backing up a missing workbook should raise the typed missing-file exception."""

    missing_path = tmp_path / "missing.xlsx"

    with pytest.raises(WorkbookFileMissingError):
        excel_io_module.backup(missing_path)

    if config.BACKUP_DIR.exists():
        assert list(config.BACKUP_DIR.rglob("*.xlsx")) == []


def test_backup_raises_for_corrupt_workbook(ema_home: Path, tmp_path: Path) -> None:
    """Backing up a corrupt workbook should raise an EMA error and create no backup."""

    corrupt_path = tmp_path / "corrupt.xlsx"
    corrupt_bytes = b"this is not a valid xlsx file"
    corrupt_path.write_bytes(corrupt_bytes)

    with pytest.raises(EmaError):
        excel_io_module.backup(corrupt_path)

    assert corrupt_path.exists()
    assert corrupt_path.read_bytes() == corrupt_bytes
    if config.BACKUP_DIR.exists():
        assert list(config.BACKUP_DIR.rglob("*.xlsx")) == []


def test_backup_never_modifies_original_workbook(
    ema_home: Path, formatted_wb: Path
) -> None:
    """Creating a backup should leave the original workbook bytes untouched."""

    original_bytes = formatted_wb.read_bytes()

    backup_path = excel_io_module.backup(formatted_wb)

    assert backup_path.exists()
    assert formatted_wb.read_bytes() == original_bytes
    assert backup_path.read_bytes() == original_bytes


def test_backup_preserves_workbook_size_and_bytes_exactly(
    ema_home: Path, table_wb: Path
) -> None:
    """The backup should match the original workbook byte-for-byte and size-for-size."""

    backup_path = excel_io_module.backup(table_wb)

    assert backup_path.exists()
    assert backup_path.stat().st_size == table_wb.stat().st_size
    assert backup_path.read_bytes() == table_wb.read_bytes()


def test_backup_returns_existing_backup_path(ema_home: Path, plain_wb: Path) -> None:
    """The returned path should point to an on-disk backup file."""

    backup_path = excel_io_module.backup(plain_wb)

    assert backup_path.exists()
    assert backup_path.is_file()
    assert backup_path.read_bytes() == plain_wb.read_bytes()


def test_append_row_plain_workbook_appends_values_at_next_row(
    plain_wb: Path,
) -> None:
    """Appending to a plain workbook should write one new row at the first empty row."""

    result = excel_io_module.append_row(
        plain_wb,
        _plain_schema(),
        {
            "Date": "2026-06-11",
            "Meal": "Snack",
            "Food": "Apple",
            "Calories": 95,
        },
    )

    assert isinstance(result, AppendResult)
    assert result.written_row == 5

    rows = _sheet_values(plain_wb, "FoodLog")
    assert rows[4] == ["2026-06-11", "Snack", "Apple", 95]


def test_append_row_formula_workbook_writes_non_formula_columns_only(
    formula_wb: Path,
) -> None:
    """Appending to a formula workbook should succeed when the formula column is omitted."""

    result = excel_io_module.append_row(
        formula_wb,
        _formula_schema(),
        {
            "Item": "Yogurt",
            "Quantity": 3,
            "Price": 2.25,
        },
    )

    assert isinstance(result, AppendResult)
    assert result.written_row == 5

    rows = _sheet_values(formula_wb, "OrderLines")
    assert rows[4][0:3] == ["Yogurt", 3, 2.25]


def test_append_row_offset_header_workbook_appends_after_existing_data(
    offset_wb: Path,
) -> None:
    """Appending to an offset-header workbook should honor the persisted data start row."""

    result = excel_io_module.append_row(
        offset_wb,
        _offset_schema(),
        {
            "Date": "2026-06-02",
            "Marker": "LDL",
            "Value": 101,
            "Unit": "mg/dL",
        },
    )

    assert isinstance(result, AppendResult)
    assert result.written_row == 7

    rows = _sheet_values(offset_wb, "Bloodwork")
    assert rows[6] == ["2026-06-02", "LDL", 101, "mg/dL"]


def test_append_row_unknown_column_raises_validation_error(plain_wb: Path) -> None:
    """Unknown input columns should fail validation instead of being ignored."""

    with pytest.raises(EmaValidationError):
        excel_io_module.append_row(
            plain_wb,
            _plain_schema(),
            {
                "Date": "2026-06-11",
                "Meal": "Snack",
                "Food": "Apple",
                "Calories": 95,
                "Mood": "Happy",
            },
        )


def test_append_row_formula_column_write_raises_validation_error(
    formula_wb: Path,
) -> None:
    """Providing a value for a formula column should be rejected."""

    with pytest.raises(EmaValidationError):
        excel_io_module.append_row(
            formula_wb,
            _formula_schema(),
            {
                "Item": "Yogurt",
                "Quantity": 3,
                "Price": 2.25,
                "Total": 6.75,
            },
        )


def test_append_row_missing_required_column_raises_validation_error(
    plain_wb: Path,
) -> None:
    """Omitting a required column should fail validation."""

    with pytest.raises(EmaValidationError):
        excel_io_module.append_row(
            plain_wb,
            _plain_schema(),
            {
                "Date": "2026-06-11",
                "Meal": "Snack",
                "Food": "Apple",
            },
        )


def test_append_row_keeps_existing_workbook_data_intact(plain_wb: Path) -> None:
    """Appending should preserve all pre-existing workbook rows unchanged."""

    original_rows = _sheet_values(plain_wb, "FoodLog")

    excel_io_module.append_row(
        plain_wb,
        _plain_schema(),
        {
            "Date": "2026-06-11",
            "Meal": "Snack",
            "Food": "Apple",
            "Calories": 95,
        },
    )

    updated_rows = _sheet_values(plain_wb, "FoodLog")
    assert updated_rows[:4] == original_rows[:4]


def test_append_row_new_row_appears_exactly_once(plain_wb: Path) -> None:
    """Appending should add exactly one matching new row."""

    new_row = ["2026-06-11", "Snack", "Apple", 95]

    excel_io_module.append_row(
        plain_wb,
        _plain_schema(),
        {
            "Date": new_row[0],
            "Meal": new_row[1],
            "Food": new_row[2],
            "Calories": new_row[3],
        },
    )

    rows = _sheet_values(plain_wb, "FoodLog")
    assert rows.count(new_row) == 1


def test_append_row_creates_backup_before_write(
    ema_home: Path, plain_wb: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """append_row should create a backup before delegating to the atomic save step."""

    events: list[str] = []
    real_atomic_save = excel_io_module.atomic_save

    def fake_backup(path: Path) -> Path:
        events.append("backup")
        backup_path = config.BACKUP_DIR / "plain" / "plain-test-backup.xlsx"
        backup_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(path, backup_path)
        return backup_path

    def checked_atomic_save(workbook: object, path: Path) -> Path:
        assert events == ["backup"]
        events.append("save")
        return real_atomic_save(workbook, path)

    monkeypatch.setattr(excel_io_module, "backup", fake_backup)
    monkeypatch.setattr(excel_io_module, "atomic_save", checked_atomic_save)

    excel_io_module.append_row(
        plain_wb,
        _plain_schema(),
        {
            "Date": "2026-06-12",
            "Meal": "Snack",
            "Food": "Apple",
            "Calories": 95,
        },
    )

    assert events == ["backup", "save"]


def test_append_row_returns_backup_path(ema_home: Path, plain_wb: Path) -> None:
    """append_row should return the created backup path in its result."""

    result = excel_io_module.append_row(
        plain_wb,
        _plain_schema(),
        {
            "Date": "2026-06-12",
            "Meal": "Snack",
            "Food": "Apple",
            "Calories": 95,
        },
    )

    assert result.backup_path is not None
    assert Path(result.backup_path).exists()


def test_append_row_returns_undo_token(ema_home: Path, plain_wb: Path) -> None:
    """append_row should expose an undo token for the created backup."""

    result = excel_io_module.append_row(
        plain_wb,
        _plain_schema(),
        {
            "Date": "2026-06-12",
            "Meal": "Snack",
            "Food": "Apple",
            "Calories": 95,
        },
    )

    assert result.undo_token is not None
    assert result.undo_token == result.backup_path


def test_append_row_invokes_verify_write(
    ema_home: Path, plain_wb: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """append_row should invoke verify_write after saving."""

    verify_called = False

    def fake_verify_write(*args: object, **kwargs: object) -> None:
        nonlocal verify_called
        verify_called = True

    monkeypatch.setattr(excel_io_module, "verify_write", fake_verify_write)

    excel_io_module.append_row(
        plain_wb,
        _plain_schema(),
        {
            "Date": "2026-06-12",
            "Meal": "Snack",
            "Food": "Apple",
            "Calories": 95,
        },
    )

    assert verify_called is True


def test_append_row_verify_failure_restores_backup_automatically(
    ema_home: Path, plain_wb: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """If verify_write fails, append_row should restore the workbook from backup."""

    original_bytes = plain_wb.read_bytes()

    def failing_verify_write(*args: object, **kwargs: object) -> None:
        raise WriteVerificationError("simulated verification failure")

    monkeypatch.setattr(excel_io_module, "verify_write", failing_verify_write)

    with pytest.raises(WriteVerificationError, match="simulated verification failure"):
        excel_io_module.append_row(
            plain_wb,
            _plain_schema(),
            {
                "Date": "2026-06-12",
                "Meal": "Snack",
                "Food": "Apple",
                "Calories": 95,
            },
        )

    assert plain_wb.read_bytes() == original_bytes

def test_append_row_dispatches_to_plain_helper_when_schema_has_no_table_name(
    plain_wb: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """append_row should route plain schemas through the plain-row helper."""

    schema = _plain_schema()
    row = {
        "Date": "2026-06-12",
        "Meal": "Snack",
        "Food": "Apple",
        "Calories": 95,
    }
    sentinel = AppendResult(
        ok=True,
        workbook_id="plain",
        sheet=schema.sheet,
        written_row=99,
        row_preview=row,
        backup_path="backup.xlsx",
        undo_token="backup.xlsx",
        dry_run=False,
        message="plain helper called",
    )
    called: dict[str, object] = {}

    def fake_plain_helper(path: Path, helper_schema: WorkbookSchema, helper_row: dict[str, object]) -> AppendResult:
        called["path"] = path
        called["schema"] = helper_schema
        called["row"] = helper_row
        return sentinel

    def fail_table_helper(*_args: object, **_kwargs: object) -> AppendResult:
        raise AssertionError("table helper should not be used for plain schemas")

    monkeypatch.setattr(excel_io_module, "_append_plain_row", fake_plain_helper, raising=False)
    monkeypatch.setattr(excel_io_module, "_append_table_row", fail_table_helper, raising=False)

    result = excel_io_module.append_row(plain_wb, schema, row)

    assert result is sentinel
    assert called == {"path": plain_wb, "schema": schema, "row": row}



def test_append_row_dispatches_to_table_helper_when_schema_has_table_name(
    table_wb: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """append_row should route table schemas through the table-row helper."""

    schema = _table_schema()
    row = {
        "Date": "2026-06-12",
        "Store": "Lidl",
        "Amount": 22.75,
        "Category": "Groceries",
    }
    sentinel = AppendResult(
        ok=True,
        workbook_id="table",
        sheet=schema.sheet,
        written_row=55,
        row_preview=row,
        backup_path="backup.xlsx",
        undo_token="backup.xlsx",
        dry_run=False,
        message="table helper called",
    )
    called: dict[str, object] = {}

    def fail_plain_helper(*_args: object, **_kwargs: object) -> AppendResult:
        raise AssertionError("plain helper should not be used for table schemas")

    def fake_table_helper(path: Path, helper_schema: WorkbookSchema, helper_row: dict[str, object]) -> AppendResult:
        called["path"] = path
        called["schema"] = helper_schema
        called["row"] = helper_row
        return sentinel

    monkeypatch.setattr(excel_io_module, "_append_plain_row", fail_plain_helper, raising=False)
    monkeypatch.setattr(excel_io_module, "_append_table_row", fake_table_helper, raising=False)

    result = excel_io_module.append_row(table_wb, schema, row)

    assert result is sentinel
    assert called == {"path": table_wb, "schema": schema, "row": row}

def test_verify_write_succeeds_when_appended_row_matches_expected_plain_workbook(
    plain_wb: Path,
) -> None:
    """Verification should pass when the target row exists and matches exactly."""

    schema = _plain_schema()
    expected = {
        "Date": "2026-06-11",
        "Meal": "Snack",
        "Food": "Apple",
        "Calories": 95,
    }
    _write_row_using_schema(plain_wb, schema, 5, expected)

    excel_io_module.verify_write(plain_wb, schema, expected, 5)


def test_verify_write_fails_when_target_row_does_not_exist(plain_wb: Path) -> None:
    """Verification should fail when the requested row is beyond the worksheet data."""

    with pytest.raises(WriteVerificationError):
        excel_io_module.verify_write(
            plain_wb,
            _plain_schema(),
            {
                "Date": "2026-06-11",
                "Meal": "Snack",
                "Food": "Apple",
                "Calories": 95,
            },
            999,
        )


def test_verify_write_fails_when_cell_values_differ(plain_wb: Path) -> None:
    """Verification should fail when any expected cell value differs from the workbook."""

    schema = _plain_schema()
    _write_row_using_schema(
        plain_wb,
        schema,
        5,
        {
            "Date": "2026-06-11",
            "Meal": "Snack",
            "Food": "Apple",
            "Calories": 95,
        },
    )

    with pytest.raises(WriteVerificationError):
        excel_io_module.verify_write(
            plain_wb,
            schema,
            {
                "Date": "2026-06-11",
                "Meal": "Snack",
                "Food": "Pear",
                "Calories": 95,
            },
            5,
        )


def test_verify_write_succeeds_for_plain_workbook(plain_wb: Path) -> None:
    """Verification should succeed for a plain workbook row written with the schema."""

    schema = _plain_schema()
    expected = {
        "Date": "2026-06-11",
        "Meal": "Dinner",
        "Food": "Rice",
        "Calories": 640,
    }
    _write_row_using_schema(plain_wb, schema, 5, expected)

    excel_io_module.verify_write(plain_wb, schema, expected, 5)


def test_verify_write_succeeds_for_formula_workbook_when_only_non_formula_columns_checked(
    formula_wb: Path,
) -> None:
    """Verification should succeed for formula workbooks when expected data omits formula columns."""

    schema = _formula_schema()
    _write_row_using_schema(
        formula_wb,
        schema,
        5,
        {
            "Item": "Yogurt",
            "Quantity": 3,
            "Price": 2.25,
            "Total": "=B5*C5",
        },
    )

    excel_io_module.verify_write(
        formula_wb,
        schema,
        {
            "Item": "Yogurt",
            "Quantity": 3,
            "Price": 2.25,
        },
        5,
    )


def test_verify_write_succeeds_for_offset_header_workbook(offset_wb: Path) -> None:
    """Verification should succeed for a workbook whose data starts below row 1."""

    schema = _offset_schema()
    expected = {
        "Date": "2026-06-02",
        "Marker": "LDL",
        "Value": 101,
        "Unit": "mg/dL",
    }
    _write_row_using_schema(offset_wb, schema, 7, expected)

    excel_io_module.verify_write(offset_wb, schema, expected, 7)


def test_verify_write_ignores_formula_columns(formula_wb: Path) -> None:
    """Verification should ignore schema columns marked as formulas."""

    schema = _formula_schema()
    _write_row_using_schema(
        formula_wb,
        schema,
        5,
        {
            "Item": "Yogurt",
            "Quantity": 3,
            "Price": 2.25,
            "Total": "=B5*C5",
        },
    )

    excel_io_module.verify_write(
        formula_wb,
        schema,
        {
            "Item": "Yogurt",
            "Quantity": 3,
            "Price": 2.25,
            "Total": 999999,
        },
        5,
    )


def test_verify_write_raises_ema_exception_on_mismatch(plain_wb: Path) -> None:
    """Mismatches should raise the typed EMA verification exception."""

    schema = _plain_schema()
    _write_row_using_schema(
        plain_wb,
        schema,
        5,
        {
            "Date": "2026-06-11",
            "Meal": "Snack",
            "Food": "Apple",
            "Calories": 95,
        },
    )

    with pytest.raises(WriteVerificationError):
        excel_io_module.verify_write(
            plain_wb,
            schema,
            {
                "Date": "2026-06-11",
                "Meal": "Snack",
                "Food": "Apple",
                "Calories": 100,
            },
            5,
        )


def test_verify_write_uses_workbook_schema_column_definitions(plain_wb: Path) -> None:
    """Verification should locate values by schema column definitions, not dict order."""

    schema = _plain_schema()
    _write_row_using_schema(
        plain_wb,
        schema,
        5,
        {
            "Date": "2026-06-11",
            "Meal": "Snack",
            "Food": "Apple",
            "Calories": 95,
        },
    )

    excel_io_module.verify_write(
        plain_wb,
        schema,
        {
            "Calories": 95,
            "Food": "Apple",
            "Meal": "Snack",
            "Date": "2026-06-11",
        },
        5,
    )


def test_verify_write_does_not_modify_workbook_contents(plain_wb: Path) -> None:
    """Verification should read workbook contents without mutating the file bytes."""

    schema = _plain_schema()
    expected = {
        "Date": "2026-06-11",
        "Meal": "Snack",
        "Food": "Apple",
        "Calories": 95,
    }
    _write_row_using_schema(plain_wb, schema, 5, expected)
    original_bytes = plain_wb.read_bytes()

    excel_io_module.verify_write(plain_wb, schema, expected, 5)

    assert plain_wb.read_bytes() == original_bytes


def test_atomic_save_successfully_replaces_workbook(plain_wb: Path) -> None:
    """Atomic save should persist workbook changes to the target path."""

    workbook = load_workbook(plain_wb)
    try:
        worksheet = workbook["FoodLog"]
        worksheet["B2"] = "Brunch"

        returned_path = excel_io_module.atomic_save(workbook, plain_wb)
    finally:
        workbook.close()

    assert returned_path == plain_wb
    assert returned_path.exists()
    assert _sheet_values(plain_wb, "FoodLog")[1][1] == "Brunch"


def test_atomic_save_resulting_workbook_remains_valid_xlsx_file(plain_wb: Path) -> None:
    """A successful atomic save should leave a readable .xlsx file on disk."""

    workbook = load_workbook(plain_wb)
    try:
        worksheet = workbook["FoodLog"]
        worksheet["D2"] = 360
        excel_io_module.atomic_save(workbook, plain_wb)
    finally:
        workbook.close()

    reopened = load_workbook(plain_wb)
    try:
        assert "FoodLog" in reopened.sheetnames
        assert reopened["FoodLog"]["D2"].value == 360
    finally:
        reopened.close()


def test_atomic_save_replaces_original_only_after_temp_save_succeeds(
    monkeypatch: pytest.MonkeyPatch, plain_wb: Path
) -> None:
    """The original workbook should remain intact until the final replace step."""

    original_bytes = plain_wb.read_bytes()
    real_replace = os.replace

    def checked_replace(src: os.PathLike[str] | str, dst: os.PathLike[str] | str) -> None:
        src_path = Path(src)
        dst_path = Path(dst)
        assert src_path.exists()
        assert dst_path.read_bytes() == original_bytes
        real_replace(src, dst)

    monkeypatch.setattr("ema.excel_io.os.replace", checked_replace)

    workbook = load_workbook(plain_wb)
    try:
        workbook["FoodLog"]["B2"] = "Brunch"
        excel_io_module.atomic_save(workbook, plain_wb)
    finally:
        workbook.close()

    assert _sheet_values(plain_wb, "FoodLog")[1][1] == "Brunch"


def test_atomic_save_removes_temporary_file_after_successful_save(plain_wb: Path) -> None:
    """Successful atomic save should not leave orphan temp files behind."""

    before_files = _sibling_file_names(plain_wb)

    workbook = load_workbook(plain_wb)
    try:
        workbook["FoodLog"]["C2"] = "Granola"
        excel_io_module.atomic_save(workbook, plain_wb)
    finally:
        workbook.close()

    after_files = _sibling_file_names(plain_wb)
    assert after_files == before_files


def test_atomic_save_removes_temporary_file_after_failed_save(
    monkeypatch: pytest.MonkeyPatch, plain_wb: Path
) -> None:
    """Failed workbook.save should not leave any temp file behind."""

    before_files = _sibling_file_names(plain_wb)
    original_bytes = plain_wb.read_bytes()

    workbook = load_workbook(plain_wb)
    monkeypatch.setattr(
        workbook,
        "save",
        lambda _path: (_ for _ in ()).throw(OSError("simulated save failure")),
    )

    try:
        with pytest.raises(OSError, match="simulated save failure"):
            excel_io_module.atomic_save(workbook, plain_wb)
    finally:
        workbook.close()

    assert plain_wb.read_bytes() == original_bytes
    assert _sibling_file_names(plain_wb) == before_files


def test_atomic_save_simulated_save_failure_leaves_original_unchanged(
    monkeypatch: pytest.MonkeyPatch, plain_wb: Path
) -> None:
    """If saving the temp workbook fails, the original workbook bytes must be preserved."""

    original_bytes = plain_wb.read_bytes()
    workbook = load_workbook(plain_wb)
    monkeypatch.setattr(
        workbook,
        "save",
        lambda _path: (_ for _ in ()).throw(OSError("simulated save failure")),
    )

    try:
        with pytest.raises(OSError, match="simulated save failure"):
            excel_io_module.atomic_save(workbook, plain_wb)
    finally:
        workbook.close()

    assert plain_wb.read_bytes() == original_bytes


def test_atomic_save_simulated_replace_failure_leaves_original_unchanged(
    monkeypatch: pytest.MonkeyPatch, plain_wb: Path
) -> None:
    """If the final replace fails, the original workbook should remain untouched."""

    before_files = _sibling_file_names(plain_wb)
    original_bytes = plain_wb.read_bytes()

    def failing_replace(_src: os.PathLike[str] | str, _dst: os.PathLike[str] | str) -> None:
        raise OSError("simulated replace failure")

    monkeypatch.setattr("ema.excel_io.os.replace", failing_replace)

    workbook = load_workbook(plain_wb)
    try:
        workbook["FoodLog"]["B2"] = "Brunch"
        with pytest.raises(OSError, match="simulated replace failure"):
            excel_io_module.atomic_save(workbook, plain_wb)
    finally:
        workbook.close()

    assert plain_wb.read_bytes() == original_bytes
    assert _sibling_file_names(plain_wb) == before_files


def test_atomic_save_preserves_workbook_contents(plain_wb: Path) -> None:
    """Atomic save should persist intended cell changes without disturbing other data."""

    original_rows = _sheet_values(plain_wb, "FoodLog")

    workbook = load_workbook(plain_wb)
    try:
        workbook["FoodLog"]["D4"] = 650
        excel_io_module.atomic_save(workbook, plain_wb)
    finally:
        workbook.close()

    updated_rows = _sheet_values(plain_wb, "FoodLog")
    assert updated_rows[0:3] == original_rows[0:3]
    assert updated_rows[3][3] == 650


def test_atomic_save_preserves_workbook_formatting_and_features(
    formatted_wb: Path,
) -> None:
    """Atomic save should preserve workbook styling, panes, widths, and chart features."""

    workbook = load_workbook(formatted_wb)
    try:
        worksheet = workbook["Metrics"]
        worksheet["B2"] = 80.0
        excel_io_module.atomic_save(workbook, formatted_wb)
    finally:
        workbook.close()

    reopened = load_workbook(formatted_wb)
    try:
        worksheet = reopened["Metrics"]
        assert worksheet["A1"].fill.fill_type == "solid"
        assert worksheet["A1"].fill.fgColor.rgb == "001F4E78"
        assert worksheet["A1"].font.bold is True
        assert worksheet.freeze_panes == "A2"
        assert worksheet.column_dimensions["A"].width == 14
        assert worksheet.column_dimensions["B"].width == 12
        assert worksheet.column_dimensions["C"].width == 12
        assert len(worksheet._charts) == 1
    finally:
        reopened.close()


def test_atomic_save_returned_workbook_path_exists_and_is_valid(plain_wb: Path) -> None:
    """The returned atomic-save path should exist and open as a valid workbook."""

    workbook = load_workbook(plain_wb)
    try:
        workbook["FoodLog"]["A4"] = "2026-06-12"
        returned_path = excel_io_module.atomic_save(workbook, plain_wb)
    finally:
        workbook.close()

    assert returned_path.exists()
    reopened = load_workbook(returned_path)
    try:
        assert "FoodLog" in reopened.sheetnames
    finally:
        reopened.close()


def test_append_row_table_appends_row_into_excel_table_workbook(table_wb: Path) -> None:
    """Appending to a table workbook should add one row of values to the sheet."""

    result = excel_io_module.append_row(
        table_wb,
        _table_schema(),
        {
            "Date": "2026-06-12",
            "Store": "Lidl",
            "Amount": 22.75,
            "Category": "Groceries",
        },
    )

    assert isinstance(result, AppendResult)
    assert result.written_row == 5
    rows = _sheet_values(table_wb, "Expenses")
    assert rows[4] == ["2026-06-12", "Lidl", 22.75, "Groceries"]


def test_append_row_table_writes_new_row_after_last_table_row(table_wb: Path) -> None:
    """The new table row should be written immediately after the current table data."""

    excel_io_module.append_row(
        table_wb,
        _table_schema(),
        {
            "Date": "2026-06-12",
            "Store": "Lidl",
            "Amount": 22.75,
            "Category": "Groceries",
        },
    )

    rows = _sheet_values(table_wb, "Expenses")
    assert rows[4] == ["2026-06-12", "Lidl", 22.75, "Groceries"]


def test_append_row_table_expands_excel_table_range_correctly(table_wb: Path) -> None:
    """Appending to a table workbook should extend the table range by one row."""

    excel_io_module.append_row(
        table_wb,
        _table_schema(),
        {
            "Date": "2026-06-12",
            "Store": "Lidl",
            "Amount": 22.75,
            "Category": "Groceries",
        },
    )

    workbook = load_workbook(table_wb)
    try:
        worksheet = workbook["Expenses"]
        assert worksheet.tables["ExpensesTable"].ref == "A1:D5"
    finally:
        workbook.close()


def test_append_row_table_keeps_existing_table_data_unchanged(table_wb: Path) -> None:
    """Appending to the table should preserve all existing data rows."""

    original_rows = _sheet_values(table_wb, "Expenses")

    excel_io_module.append_row(
        table_wb,
        _table_schema(),
        {
            "Date": "2026-06-12",
            "Store": "Lidl",
            "Amount": 22.75,
            "Category": "Groceries",
        },
    )

    updated_rows = _sheet_values(table_wb, "Expenses")
    assert updated_rows[:4] == original_rows[:4]


def test_append_row_table_new_row_appears_exactly_once(table_wb: Path) -> None:
    """Appending to the table should add exactly one matching new row."""

    new_row = ["2026-06-12", "Lidl", 22.75, "Groceries"]

    excel_io_module.append_row(
        table_wb,
        _table_schema(),
        {
            "Date": new_row[0],
            "Store": new_row[1],
            "Amount": new_row[2],
            "Category": new_row[3],
        },
    )

    rows = _sheet_values(table_wb, "Expenses")
    assert rows.count(new_row) == 1


def test_append_row_table_formula_columns_are_protected(table_wb: Path) -> None:
    """Table append should reject writes to schema columns marked as formulas."""

    schema = _table_schema()
    schema.columns[2].is_formula = True

    with pytest.raises(EmaValidationError):
        excel_io_module.append_row(
            table_wb,
            schema,
            {
                "Date": "2026-06-12",
                "Store": "Lidl",
                "Amount": 22.75,
                "Category": "Groceries",
            },
        )


def test_append_row_table_unknown_columns_raise_validation_error(table_wb: Path) -> None:
    """Unknown input columns should fail validation for table appends."""

    with pytest.raises(EmaValidationError):
        excel_io_module.append_row(
            table_wb,
            _table_schema(),
            {
                "Date": "2026-06-12",
                "Store": "Lidl",
                "Amount": 22.75,
                "Category": "Groceries",
                "Currency": "EUR",
            },
        )


def test_append_row_table_missing_required_columns_raise_validation_error(
    table_wb: Path,
) -> None:
    """Missing required columns should fail validation for table appends."""

    with pytest.raises(EmaValidationError):
        excel_io_module.append_row(
            table_wb,
            _table_schema(),
            {
                "Date": "2026-06-12",
                "Store": "Lidl",
                "Amount": 22.75,
            },
        )


def test_append_row_table_preserves_valid_xlsx_workbook(table_wb: Path) -> None:
    """After table append, the workbook should still open as a valid .xlsx file."""

    excel_io_module.append_row(
        table_wb,
        _table_schema(),
        {
            "Date": "2026-06-12",
            "Store": "Lidl",
            "Amount": 22.75,
            "Category": "Groceries",
        },
    )

    workbook = load_workbook(table_wb)
    try:
        assert "Expenses" in workbook.sheetnames
        assert workbook["Expenses"]["B5"].value == "Lidl"
    finally:
        workbook.close()


def test_append_row_table_preserves_table_formatting(table_wb: Path) -> None:
    """Appending should preserve the Excel Table's style metadata."""

    excel_io_module.append_row(
        table_wb,
        _table_schema(),
        {
            "Date": "2026-06-12",
            "Store": "Lidl",
            "Amount": 22.75,
            "Category": "Groceries",
        },
    )

    workbook = load_workbook(table_wb)
    try:
        table = workbook["Expenses"].tables["ExpensesTable"]
        assert table.tableStyleInfo is not None
        assert table.tableStyleInfo.name == "TableStyleMedium2"
        assert table.tableStyleInfo.showRowStripes is True
    finally:
        workbook.close()


def test_append_row_table_creates_backup_before_write(
    ema_home: Path, table_wb: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """append_row should create a backup before delegating to the atomic save step for tables."""

    events: list[str] = []
    real_atomic_save = excel_io_module.atomic_save

    def fake_backup(path: Path) -> Path:
        events.append("backup")
        backup_path = config.BACKUP_DIR / "table" / "table-test-backup.xlsx"
        backup_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(path, backup_path)
        return backup_path

    def checked_atomic_save(workbook: object, path: Path) -> Path:
        assert events == ["backup"]
        events.append("save")
        return real_atomic_save(workbook, path)

    monkeypatch.setattr(excel_io_module, "backup", fake_backup)
    monkeypatch.setattr(excel_io_module, "atomic_save", checked_atomic_save)

    excel_io_module.append_row(
        table_wb,
        _table_schema(),
        {
            "Date": "2026-06-12",
            "Store": "Lidl",
            "Amount": 22.75,
            "Category": "Groceries",
        },
    )

    assert events == ["backup", "save"]


def test_append_row_table_returns_backup_path(ema_home: Path, table_wb: Path) -> None:
    """append_row should return the created backup path in its result for tables."""

    result = excel_io_module.append_row(
        table_wb,
        _table_schema(),
        {
            "Date": "2026-06-12",
            "Store": "Lidl",
            "Amount": 22.75,
            "Category": "Groceries",
        },
    )

    assert result.backup_path is not None
    assert Path(result.backup_path).exists()


def test_append_row_table_returns_undo_token(ema_home: Path, table_wb: Path) -> None:
    """append_row should expose an undo token for the created backup for tables."""

    result = excel_io_module.append_row(
        table_wb,
        _table_schema(),
        {
            "Date": "2026-06-12",
            "Store": "Lidl",
            "Amount": 22.75,
            "Category": "Groceries",
        },
    )

    assert result.undo_token is not None
    assert result.undo_token == result.backup_path


def test_append_row_table_invokes_verify_write(
    ema_home: Path, table_wb: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """append_row should invoke verify_write after saving for tables."""

    verify_called = False

    def fake_verify_write(*args: object, **kwargs: object) -> None:
        nonlocal verify_called
        verify_called = True

    monkeypatch.setattr(excel_io_module, "verify_write", fake_verify_write)

    excel_io_module.append_row(
        table_wb,
        _table_schema(),
        {
            "Date": "2026-06-12",
            "Store": "Lidl",
            "Amount": 22.75,
            "Category": "Groceries",
        },
    )

    assert verify_called is True


def test_append_row_table_verify_failure_restores_backup_automatically(
    ema_home: Path, table_wb: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """If verify_write fails, append_row should restore the workbook from backup for tables."""

    original_bytes = table_wb.read_bytes()

    def failing_verify_write(*args: object, **kwargs: object) -> None:
        raise WriteVerificationError("simulated verification failure")

    monkeypatch.setattr(excel_io_module, "verify_write", failing_verify_write)

    with pytest.raises(WriteVerificationError, match="simulated verification failure"):
        excel_io_module.append_row(
            table_wb,
            _table_schema(),
            {
                "Date": "2026-06-12",
                "Store": "Lidl",
                "Amount": 22.75,
                "Category": "Groceries",
            },
        )

    assert table_wb.read_bytes() == original_bytes


def test_undo_last_restores_workbook_to_exact_backup_state(plain_wb: Path) -> None:
    """Undo should restore the workbook bytes exactly to the chosen backup state."""

    original_bytes = plain_wb.read_bytes()
    backup_path = excel_io_module.backup(plain_wb)
    excel_io_module.append_row(
        plain_wb,
        _plain_schema(),
        {
            "Date": "2026-06-12",
            "Meal": "Snack",
            "Food": "Apple",
            "Calories": 95,
        },
    )

    result = excel_io_module.undo_last(plain_wb, str(backup_path))

    assert isinstance(result, UndoResult)
    assert plain_wb.read_bytes() == original_bytes


def test_undo_last_removes_last_appended_row(plain_wb: Path) -> None:
    """Undo should remove the last row added after the backup was taken."""

    backup_path = excel_io_module.backup(plain_wb)
    excel_io_module.append_row(
        plain_wb,
        _plain_schema(),
        {
            "Date": "2026-06-12",
            "Meal": "Snack",
            "Food": "Apple",
            "Calories": 95,
        },
    )

    excel_io_module.undo_last(plain_wb, str(backup_path))

    rows = _sheet_values(plain_wb, "FoodLog")
    assert rows == [
        ["Date", "Meal", "Food", "Calories"],
        ["2026-06-09", "Breakfast", "Oats", 350],
        ["2026-06-09", "Lunch", "Chicken Salad", 520],
        ["2026-06-10", "Dinner", "Pasta", 640],
    ]


def test_undo_last_restores_excel_table_workbook_correctly(table_wb: Path) -> None:
    """Undo should restore the pre-append table workbook bytes and table range."""

    original_bytes = table_wb.read_bytes()
    backup_path = excel_io_module.backup(table_wb)
    excel_io_module.append_row(
        table_wb,
        _table_schema(),
        {
            "Date": "2026-06-12",
            "Store": "Lidl",
            "Amount": 22.75,
            "Category": "Groceries",
        },
    )

    excel_io_module.undo_last(table_wb, str(backup_path))

    assert table_wb.read_bytes() == original_bytes
    workbook = load_workbook(table_wb)
    try:
        assert workbook["Expenses"].tables["ExpensesTable"].ref == "A1:D4"
    finally:
        workbook.close()


def test_undo_last_restores_workbook_formatting_and_features(
    formatted_wb: Path,
) -> None:
    """Undo should restore styled workbooks without losing formatting features."""

    original_bytes = formatted_wb.read_bytes()
    backup_path = excel_io_module.backup(formatted_wb)

    workbook = load_workbook(formatted_wb)
    try:
        worksheet = workbook["Metrics"]
        worksheet["B2"] = 80.0
        excel_io_module.atomic_save(workbook, formatted_wb)
    finally:
        workbook.close()

    excel_io_module.undo_last(formatted_wb, str(backup_path))

    assert formatted_wb.read_bytes() == original_bytes
    restored = load_workbook(formatted_wb)
    try:
        worksheet = restored["Metrics"]
        assert worksheet["A1"].fill.fill_type == "solid"
        assert worksheet["A1"].fill.fgColor.rgb == "001F4E78"
        assert worksheet["A1"].font.bold is True
        assert worksheet.freeze_panes == "A2"
        assert worksheet.column_dimensions["A"].width == 14
        assert len(worksheet._charts) == 1
    finally:
        restored.close()


def test_undo_last_restores_formula_workbooks_correctly(formula_wb: Path) -> None:
    """Undo should restore formula workbooks to their exact pre-append bytes."""

    original_bytes = formula_wb.read_bytes()
    backup_path = excel_io_module.backup(formula_wb)
    excel_io_module.append_row(
        formula_wb,
        _formula_schema(),
        {
            "Item": "Yogurt",
            "Quantity": 3,
            "Price": 2.25,
        },
    )

    excel_io_module.undo_last(formula_wb, str(backup_path))

    assert formula_wb.read_bytes() == original_bytes


def test_undo_last_missing_backup_file_raises_typed_exception(plain_wb: Path) -> None:
    """Undo should fail with a typed EMA exception when the backup file is missing."""

    missing_backup = plain_wb.parent / "missing-backup.xlsx"

    with pytest.raises(UndoError):
        excel_io_module.undo_last(plain_wb, str(missing_backup))


def test_undo_last_invalid_token_raises_typed_exception(plain_wb: Path) -> None:
    """Undo should reject tokens that do not resolve to a valid backup path."""

    with pytest.raises(UndoError):
        excel_io_module.undo_last(plain_wb, "not-a-real-backup-token")


def test_undo_last_does_not_affect_unrelated_workbooks(
    plain_wb: Path, table_wb: Path
) -> None:
    """Undo should only restore the targeted workbook and leave others untouched."""

    unrelated_before = table_wb.read_bytes()
    backup_path = excel_io_module.backup(plain_wb)
    excel_io_module.append_row(
        plain_wb,
        _plain_schema(),
        {
            "Date": "2026-06-12",
            "Meal": "Snack",
            "Food": "Apple",
            "Calories": 95,
        },
    )

    excel_io_module.undo_last(plain_wb, str(backup_path))

    assert table_wb.read_bytes() == unrelated_before


def test_undo_last_restored_workbook_remains_valid_xlsx_file(plain_wb: Path) -> None:
    """After undo, the restored workbook should still open as a valid .xlsx file."""

    backup_path = excel_io_module.backup(plain_wb)
    excel_io_module.append_row(
        plain_wb,
        _plain_schema(),
        {
            "Date": "2026-06-12",
            "Meal": "Snack",
            "Food": "Apple",
            "Calories": 95,
        },
    )

    returned = excel_io_module.undo_last(plain_wb, str(backup_path))

    workbook = load_workbook(returned.restored_from if isinstance(returned, UndoResult) else plain_wb)
    workbook.close()
    restored = load_workbook(plain_wb)
    restored.close()


def test_undo_last_with_multiple_backups_restores_the_correct_version(plain_wb: Path) -> None:
    """Undo should restore whichever backup token is provided, even across versions."""

    original_bytes = plain_wb.read_bytes()
    first_backup = excel_io_module.backup(plain_wb)

    excel_io_module.append_row(
        plain_wb,
        _plain_schema(),
        {
            "Date": "2026-06-12",
            "Meal": "Snack",
            "Food": "Apple",
            "Calories": 95,
        },
    )
    one_append_bytes = plain_wb.read_bytes()
    second_backup = excel_io_module.backup(plain_wb)

    excel_io_module.append_row(
        plain_wb,
        _plain_schema(),
        {
            "Date": "2026-06-13",
            "Meal": "Lunch",
            "Food": "Wrap",
            "Calories": 430,
        },
    )

    excel_io_module.undo_last(plain_wb, str(second_backup))
    assert plain_wb.read_bytes() == one_append_bytes

    excel_io_module.undo_last(plain_wb, str(first_backup))
    assert plain_wb.read_bytes() == original_bytes
