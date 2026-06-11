"""Workbook fixture generators for EMA tests.

These helpers build real `.xlsx` files on disk so later milestones can test
schema discovery and safe writing against concrete workbooks rather than mocks.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


def _new_workbook(sheet_name: str) -> tuple[Workbook, object]:
    """Create a workbook with a single renamed active sheet."""

    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = sheet_name
    return workbook, worksheet


def make_plain(path: Path) -> Path:
    """Create a simple workbook with headers on row 1 and sample rows."""

    workbook, worksheet = _new_workbook("FoodLog")

    rows = [
        ["Date", "Meal", "Food", "Calories"],
        ["2026-06-09", "Breakfast", "Oats", 350],
        ["2026-06-09", "Lunch", "Chicken Salad", 520],
        ["2026-06-10", "Dinner", "Pasta", 640],
    ]

    for row in rows:
        worksheet.append(row)

    workbook.save(path)
    return path


def make_table(path: Path) -> Path:
    """Create a workbook whose data lives inside an Excel Table object."""

    workbook, worksheet = _new_workbook("Expenses")

    rows = [
        ["Date", "Store", "Amount", "Category"],
        ["2026-06-09", "Aldi", 18.50, "Groceries"],
        ["2026-06-10", "Shell", 62.10, "Fuel"],
        ["2026-06-11", "Pharmacy", 9.99, "Health"],
    ]

    for row in rows:
        worksheet.append(row)

    table = Table(displayName="ExpensesTable", ref="A1:D4")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)

    workbook.save(path)
    return path


def make_formula(path: Path) -> Path:
    """Create a workbook that includes a formula column."""

    workbook, worksheet = _new_workbook("OrderLines")

    rows = [
        ["Item", "Quantity", "Price", "Total"],
        ["Eggs", 12, 0.35],
        ["Skyr", 2, 1.75],
        ["Edamame", 1, 3.40],
    ]

    worksheet.append(rows[0])
    for row_index, row in enumerate(rows[1:], start=2):
        worksheet.append(row)
        worksheet[f"D{row_index}"] = f"=B{row_index}*C{row_index}"

    workbook.save(path)
    return path


def make_offset_headers(path: Path) -> Path:
    """Create a workbook whose headers start below title and metadata rows."""

    workbook, worksheet = _new_workbook("Bloodwork")

    worksheet["A1"] = "Quarterly Lab Results"
    worksheet["A2"] = "Generated for fixture testing"

    rows = [
        ["Date", "Marker", "Value", "Unit"],
        ["2026-06-01", "Glucose", 92, "mg/dL"],
        ["2026-06-01", "HDL", 58, "mg/dL"],
        ["2026-06-01", "Triglycerides", 110, "mg/dL"],
    ]

    for row in rows:
        worksheet.append(row)

    workbook.save(path)
    return path


def make_formatted(path: Path) -> Path:
    """Create a styled workbook with formatting, frozen panes, and a chart."""

    workbook, worksheet = _new_workbook("Metrics")

    rows = [
        ["Week", "Weight", "Protein"],
        ["2026-W22", 79.8, 145],
        ["2026-W23", 79.2, 152],
        ["2026-W24", 78.9, 149],
    ]

    for row in rows:
        worksheet.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 14
    worksheet.column_dimensions["B"].width = 12
    worksheet.column_dimensions["C"].width = 12

    chart = BarChart()
    chart.title = "Weekly Protein Intake"
    chart.y_axis.title = "Grams"
    chart.x_axis.title = "Week"
    data = Reference(worksheet, min_col=3, min_row=1, max_row=4)
    categories = Reference(worksheet, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    worksheet.add_chart(chart, "E2")

    workbook.save(path)
    return path


__all__ = [
    "make_formula",
    "make_formatted",
    "make_offset_headers",
    "make_plain",
    "make_table",
]
