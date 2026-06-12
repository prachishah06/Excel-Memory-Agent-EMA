"""Strict TDD tests for the M6 EMA service layer."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

import ema.config as config
import ema.service as service_module
from ema.errors import SchemaMismatchError, ValidationError as EmaValidationError, WorkbookNotFoundError
from ema.models import AppendRequest, SchemaProposal, WorkbookEntry, WorkbookSchema
from ema.registry import RegistryStore


def _make_service() -> object:
    """Build a service instance bound to the test registry path."""

    return service_module.EmaService(store=RegistryStore(config.REGISTRY_PATH))


def _sheet_values(path: Path, sheet_name: str) -> list[list[object | None]]:
    """Return worksheet values as a list of row lists."""

    workbook = load_workbook(path)
    try:
        worksheet = workbook[sheet_name]
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def test_register_workbook_confirm_false_returns_schema_proposal_without_persisting(
    ema_home: Path, plain_wb: Path
) -> None:
    """confirm=False should preview the schema and leave the registry untouched."""

    service = _make_service()

    proposal = service.register_workbook(str(plain_wb), confirm=False)

    assert isinstance(proposal, SchemaProposal)
    assert proposal.sheet == "FoodLog"
    assert proposal.table_name is None
    assert [column.name for column in proposal.columns] == [
        "Date",
        "Meal",
        "Food",
        "Calories",
    ]
    assert RegistryStore(config.REGISTRY_PATH).list() == []


def test_register_workbook_confirm_true_persists_workbook_and_captures_schema(
    ema_home: Path, table_wb: Path
) -> None:
    """confirm=True should persist a workbook entry built from the proposed schema."""

    service = _make_service()

    entry = service.register_workbook(str(table_wb), name="Business Expenses", confirm=True)

    assert isinstance(entry, WorkbookEntry)
    assert entry.name == "Business Expenses"
    assert entry.path == table_wb.resolve()
    assert entry.primary_sheet == "Expenses"
    assert entry.schema.table_name == "ExpensesTable"
    assert [column.name for column in entry.schema.columns] == [
        "Date",
        "Store",
        "Amount",
        "Category",
    ]

    persisted = RegistryStore(config.REGISTRY_PATH).get(entry.id)
    assert persisted.model_dump(mode="python") == entry.model_dump(mode="python")


def test_list_workbooks_returns_registered_workbooks(
    ema_home: Path, plain_wb: Path, table_wb: Path
) -> None:
    """The service should return all persisted workbook entries."""

    service = _make_service()
    first = service.register_workbook(str(plain_wb), name="Plain Book", confirm=True)
    second = service.register_workbook(str(table_wb), name="Table Book", confirm=True)

    entries = service.list_workbooks()

    assert [entry.id for entry in entries] == [first.id, second.id]
    assert [entry.name for entry in entries] == ["Plain Book", "Table Book"]


def test_get_workbook_schema_returns_persisted_schema(
    ema_home: Path, formula_wb: Path
) -> None:
    """The service should return the schema persisted during registration."""

    service = _make_service()
    entry = service.register_workbook(str(formula_wb), confirm=True)

    schema = service.get_workbook_schema(entry.id)

    assert isinstance(schema, WorkbookSchema)
    assert schema.model_dump(mode="python") == entry.schema.model_dump(mode="python")
    assert schema.columns[3].is_formula is True


def test_append_row_plain_workbook_uses_registered_schema_and_updates_registry(
    ema_home: Path, plain_wb: Path
) -> None:
    """Appending through the service should use the persisted schema and update bookkeeping."""

    service = _make_service()
    entry = service.register_workbook(str(plain_wb), confirm=True)

    result = service.append_row(
        AppendRequest(
            workbook_id=entry.id,
            values={
                "Date": "2026-06-12",
                "Meal": "Snack",
                "Food": "Apple",
                "Calories": 95,
            },
        )
    )

    assert result.ok is True
    assert result.sheet == entry.primary_sheet
    assert result.written_row == 5
    assert result.undo_token is not None
    rows = _sheet_values(plain_wb, "FoodLog")
    assert rows[4] == ["2026-06-12", "Snack", "Apple", 95]

    persisted = RegistryStore(config.REGISTRY_PATH).get(entry.id)
    assert persisted.last_undo_token == result.undo_token
    assert persisted.updated_at > entry.updated_at


def test_append_row_table_workbook_appends_without_callers_knowing_about_tables(
    ema_home: Path, table_wb: Path
) -> None:
    """The service should use the same append API for table-backed workbooks."""

    service = _make_service()
    entry = service.register_workbook(str(table_wb), confirm=True)

    result = service.append_row(
        AppendRequest(
            workbook_id=entry.id,
            values={
                "Date": "2026-06-12",
                "Store": "Lidl",
                "Amount": 22.75,
                "Category": "Groceries",
            },
        )
    )

    assert result.ok is True
    assert result.written_row == 5
    rows = _sheet_values(table_wb, "Expenses")
    assert rows[4] == ["2026-06-12", "Lidl", 22.75, "Groceries"]

    workbook = load_workbook(table_wb)
    try:
        assert workbook["Expenses"].tables["ExpensesTable"].ref == "A1:D5"
    finally:
        workbook.close()


def test_append_row_unknown_workbook_id_raises_typed_error(ema_home: Path) -> None:
    """Unknown workbook IDs should surface the registry's typed not-found error."""

    service = _make_service()

    with pytest.raises(WorkbookNotFoundError):
        service.append_row(
            AppendRequest(
                workbook_id="missing-workbook",
                values={"Date": "2026-06-12"},
            )
        )


def test_undo_last_append_restores_previous_workbook_state(
    ema_home: Path, plain_wb: Path
) -> None:
    """Undo should restore the last append using the token tracked in the registry."""

    service = _make_service()
    entry = service.register_workbook(str(plain_wb), confirm=True)
    original_rows = _sheet_values(plain_wb, "FoodLog")

    append_result = service.append_row(
        AppendRequest(
            workbook_id=entry.id,
            values={
                "Date": "2026-06-12",
                "Meal": "Snack",
                "Food": "Apple",
                "Calories": 95,
            },
        )
    )
    assert append_result.undo_token is not None

    undo_result = service.undo_last_append(entry.id)

    assert undo_result.ok is True
    assert _sheet_values(plain_wb, "FoodLog") == original_rows


def test_service_surfaces_writer_layer_exceptions_cleanly(
    ema_home: Path, plain_wb: Path
) -> None:
    """Writer-layer validation errors should be raised unchanged by the service."""

    service = _make_service()
    entry = service.register_workbook(str(plain_wb), confirm=True)

    with pytest.raises(EmaValidationError):
        service.append_row(
            AppendRequest(
                workbook_id=entry.id,
                values={
                    "Date": "2026-06-12",
                    "Meal": "Snack",
                    "Food": "Apple",
                    "Calories": 95,
                    "Mood": "Happy",
                },
            )
        )


def test_registry_persistence_survives_new_service_instance(
    ema_home: Path, offset_wb: Path
) -> None:
    """A second service instance should load workbook state from the same registry."""

    service_one = _make_service()
    entry = service_one.register_workbook(
        str(offset_wb),
        name="Bloodwork",
        sheet="Bloodwork",
        header_row=3,
        confirm=True,
    )

    service_two = _make_service()
    listed = service_two.list_workbooks()
    loaded_schema = service_two.get_workbook_schema(entry.id)

    assert [item.id for item in listed] == [entry.id]
    assert loaded_schema.model_dump(mode="python") == entry.schema.model_dump(mode="python")


def test_append_row_surfaces_schema_drift_correctly(
    ema_home: Path, plain_wb: Path
) -> None:
    """Appending through the service should fail when the live workbook drifts from its schema."""

    service = _make_service()
    entry = service.register_workbook(str(plain_wb), confirm=True)

    workbook = load_workbook(plain_wb)
    try:
        worksheet = workbook["FoodLog"]
        worksheet["B1"] = "MealRenamed"
        workbook.save(plain_wb)
    finally:
        workbook.close()

    with pytest.raises(SchemaMismatchError):
        service.append_row(
            AppendRequest(
                workbook_id=entry.id,
                values={
                    "Date": "2026-06-12",
                    "Meal": "Snack",
                    "Food": "Apple",
                    "Calories": 95,
                },
            )
        )
