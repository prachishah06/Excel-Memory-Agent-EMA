"""Structured service layer for Excel Memory Agent."""

from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

from ema import config, excel_io, schema as schema_module
from ema.errors import UndoError
from ema.models import AppendRequest, AppendResult, SchemaProposal, UndoResult, WorkbookEntry, WorkbookSchema
from ema.registry import RegistryStore


class EmaService:
    """Thin orchestration layer over the registry, schema, and writer modules."""

    def __init__(
        self,
        store: RegistryStore | None = None,
        provider=None,
    ) -> None:
        """Initialize the service with injectable persistence and later LLM provider."""

        self._store = store or RegistryStore()
        self._provider = provider

    def register_workbook(
        self,
        path,
        name=None,
        sheet=None,
        header_row=None,
        confirm=True,
    ) -> WorkbookEntry | SchemaProposal:
        """Propose or persist workbook registration metadata."""

        workbook_path = Path(path).resolve()
        target_sheet = sheet or _default_sheet_name(workbook_path)
        proposal = schema_module.propose_schema(workbook_path, target_sheet, header_row)

        if not confirm:
            return proposal

        now = datetime.now(UTC)
        entry = WorkbookEntry(
            id=workbook_path.stem,
            name=name or workbook_path.stem,
            path=workbook_path,
            primary_sheet=target_sheet,
            schema=_persisted_schema_from_proposal(proposal),
            created_at=now,
            updated_at=now,
            last_undo_token=None,
        )
        self._store.upsert(entry)
        return entry

    def list_workbooks(self) -> list[WorkbookEntry]:
        """Return all registered workbooks from the registry."""

        return self._store.list()

    def get_workbook_schema(
        self,
        workbook_id: str,
    ) -> WorkbookSchema:
        """Return the persisted schema for one registered workbook."""

        return self._store.get(workbook_id).schema

    def append_row(
        self,
        req: AppendRequest,
    ) -> AppendResult:
        """Resolve a registered workbook and delegate one structured append."""

        entry = self._store.get(req.workbook_id)
        target_sheet = req.sheet or entry.primary_sheet
        target_schema = (
            entry.schema
            if target_sheet == entry.schema.sheet
            else entry.schema.model_copy(update={"sheet": target_sheet})
        )

        result = excel_io.append_row(entry.path, target_schema, req.values)
        updated_entry = entry.model_copy(
            update={
                "updated_at": datetime.now(UTC),
                "last_undo_token": result.undo_token,
            }
        )
        self._store.upsert(updated_entry)
        return result.model_copy(update={"workbook_id": entry.id, "sheet": target_sheet})

    def undo_last_append(
        self,
        workbook_id: str,
    ) -> UndoResult:
        """Undo the most recent append tracked for a registered workbook."""

        entry = self._store.get(workbook_id)
        if entry.last_undo_token is None:
            raise UndoError(f"No undo token available for workbook: {workbook_id}")

        result = excel_io.undo_last(entry.path, entry.last_undo_token)
        return result.model_copy(update={"workbook_id": entry.id})


def _persisted_schema_from_proposal(proposal: SchemaProposal) -> WorkbookSchema:
    """Convert a transient schema proposal into a persisted workbook schema."""

    return WorkbookSchema(
        schema_version=config.SCHEMA_VERSION,
        sheet=proposal.sheet,
        header_row=proposal.header_row,
        table_name=proposal.table_name,
        columns=proposal.columns,
        data_start_row=proposal.header_row + 1,
    )


def _default_sheet_name(path: Path) -> str:
    """Return the first worksheet name from a workbook."""

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return workbook.sheetnames[0]
    finally:
        workbook.close()


__all__ = ["EmaService"]
