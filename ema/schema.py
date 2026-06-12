"""Workbook schema discovery and validation for Excel Memory Agent.

This module is the single source of truth for workbook columns, positions,
basic type inference, and the extraction contract used by later LLM flows.
The implementation here is intentionally minimal and focused on the tested
M4 behavior.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from ema import config
from ema.errors import SchemaMismatchError, SheetNotFoundError
from ema.models import ColumnDef, ColumnType, SchemaProposal, WorkbookSchema

_TYPE_SAMPLE_LIMIT = 10


def propose_schema(path: Path, sheet: str, header_row: int | None = None) -> SchemaProposal:
    """Inspect a workbook and propose a schema for one worksheet.

    When a sheet contains exactly one Excel Table, the table range and header
    row are used automatically. Otherwise, `header_row` is used, defaulting to
    row 1.
    """

    workbook = load_workbook(path, data_only=False)
    try:
        worksheet = _get_worksheet(workbook, sheet)
        warnings: list[str] = []

        tables = _detect_tables(worksheet)
        if len(tables) == 1:
            table = tables[0]
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            detected_header_row = min_row
            headers = _read_header_row(
                worksheet,
                detected_header_row,
                min_col=min_col,
                max_col=max_col,
            )
            data_rows = range(detected_header_row + 1, max_row + 1)
            table_name = table.name
        else:
            if len(tables) > 1:
                warnings.append("Multiple Excel tables detected; using explicit header row.")

            detected_header_row = header_row or 1
            headers = _read_header_row(worksheet, detected_header_row)
            data_rows = range(detected_header_row + 1, worksheet.max_row + 1)
            table_name = None

        warnings.extend(_header_warnings(headers))
        columns = _build_columns(
            worksheet,
            headers=headers,
            data_rows=data_rows,
            header_row=detected_header_row,
        )
        warnings.extend(_formula_warnings(columns))

        if not any(
            worksheet.cell(row=row_index, column=headers[0][0]).value is not None
            for row_index in data_rows
        ) and headers:
            warnings.append("No data rows found below the detected header row.")

        return SchemaProposal(
            sheet=sheet,
            header_row=detected_header_row,
            table_name=table_name,
            columns=columns,
            warnings=warnings,
        )
    finally:
        workbook.close()


def validate_live_schema(path: Path, schema: WorkbookSchema) -> None:
    """Confirm that a workbook still matches a previously persisted schema."""

    workbook = load_workbook(path, data_only=False)
    try:
        worksheet = _get_worksheet(workbook, schema.sheet)

        if schema.table_name is not None:
            tables = _detect_tables(worksheet)
            table_by_name = {table.name: table for table in tables}
            if schema.table_name not in table_by_name:
                raise SchemaMismatchError(
                    f"Expected table '{schema.table_name}' was not found in sheet '{schema.sheet}'."
                )

            table = table_by_name[schema.table_name]
            min_col, min_row, max_col, _ = range_boundaries(table.ref)
            live_header_row = min_row
            headers = _read_header_row(
                worksheet,
                live_header_row,
                min_col=min_col,
                max_col=max_col,
            )
        else:
            live_header_row = schema.header_row
            headers = _read_header_row(worksheet, live_header_row)

        expected_headers = [column.name for column in schema.columns]
        live_headers = [header for _, header in headers]

        if live_header_row != schema.header_row:
            raise SchemaMismatchError(
                f"Header row drift detected for sheet '{schema.sheet}': "
                f"expected {schema.header_row}, found {live_header_row}."
            )

        if live_headers != expected_headers:
            raise SchemaMismatchError(
                f"Header drift detected for sheet '{schema.sheet}': "
                f"expected {expected_headers}, found {live_headers}."
            )
    finally:
        workbook.close()


def as_extraction_contract(schema: WorkbookSchema) -> dict:
    """Build a JSON-schema-like extraction contract from a persisted schema."""

    properties: dict[str, dict[str, str]] = {}
    required: list[str] = []

    for column in schema.columns:
        property_schema: dict[str, str] = {"type": _json_type_for_column(column.type)}
        if column.description is not None:
            property_schema["description"] = column.description
        properties[column.name] = property_schema

        if column.required:
            required.append(column.name)

    return {
        "type": "object",
        "additionalProperties": False,
        "properties": properties,
        "required": required,
    }


def _get_worksheet(workbook: object, sheet_name: str) -> Worksheet:
    """Return a worksheet by name or raise a typed error."""

    try:
        worksheet = workbook[sheet_name]
    except KeyError as exc:
        raise SheetNotFoundError(f"Worksheet not found: {sheet_name}") from exc

    return worksheet


def _detect_tables(worksheet: Worksheet) -> list[object]:
    """Return Excel Table objects defined on a worksheet."""

    return list(worksheet.tables.values())


def _read_header_row(
    worksheet: Worksheet,
    header_row: int,
    *,
    min_col: int | None = None,
    max_col: int | None = None,
) -> list[tuple[int, str]]:
    """Read a header row and return `(column_index, header_text)` pairs."""

    if min_col is None or max_col is None:
        values = [worksheet.cell(row=header_row, column=index).value for index in range(1, worksheet.max_column + 1)]
        non_empty_indices = [index for index, value in enumerate(values, start=1) if value is not None]
        if not non_empty_indices:
            return []
        min_col = min(non_empty_indices)
        max_col = max(non_empty_indices)

    headers: list[tuple[int, str]] = []
    for column_number in range(min_col, max_col + 1):
        value = worksheet.cell(row=header_row, column=column_number).value
        header_text = "" if value is None else str(value)
        headers.append((column_number, header_text))

    return headers


def _build_columns(
    worksheet: Worksheet,
    *,
    headers: list[tuple[int, str]],
    data_rows: range,
    header_row: int,
) -> list[ColumnDef]:
    """Create `ColumnDef` instances from header cells and sampled data."""

    columns: list[ColumnDef] = []
    for zero_based_index, (column_number, header_text) in enumerate(headers):
        sampled_values = _sample_column_values(
            worksheet,
            column_number=column_number,
            data_rows=data_rows,
            limit=_TYPE_SAMPLE_LIMIT,
        )
        has_formula = any(
            _cell_is_formula(worksheet.cell(row=row_index, column=column_number))
            for row_index in data_rows
            if row_index > header_row
        )
        columns.append(
            ColumnDef(
                name=header_text,
                index=zero_based_index,
                type=_infer_column_type(sampled_values),
                is_formula=has_formula,
            )
        )

    return columns


def _sample_column_values(
    worksheet: Worksheet,
    *,
    column_number: int,
    data_rows: range,
    limit: int,
) -> list[object]:
    """Collect up to `limit` non-empty non-formula sample values from one column."""

    samples: list[object] = []
    for row_index in data_rows:
        cell = worksheet.cell(row=row_index, column=column_number)
        if _cell_is_formula(cell):
            continue
        value = cell.value
        if value is None:
            continue
        samples.append(value)
        if len(samples) >= limit:
            break

    return samples


def _infer_column_type(samples: list[object]) -> ColumnType:
    """Infer a simple logical type from sampled workbook values."""

    if not samples:
        return ColumnType.TEXT

    if all(isinstance(value, bool) for value in samples):
        return ColumnType.BOOL

    if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in samples):
        return ColumnType.NUMBER

    if all(isinstance(value, (datetime, date)) for value in samples):
        return ColumnType.DATE

    return ColumnType.TEXT


def _cell_is_formula(cell: Cell) -> bool:
    """Return whether a workbook cell stores an Excel formula."""

    return isinstance(cell.value, str) and cell.value.startswith("=")


def _header_warnings(headers: list[tuple[int, str]]) -> list[str]:
    """Generate warnings for empty or duplicate header names."""

    warnings: list[str] = []
    header_names = [header for _, header in headers]

    if any(header == "" for header in header_names):
        warnings.append("One or more headers are empty.")

    duplicates = {header for header in header_names if header and header_names.count(header) > 1}
    if duplicates:
        warnings.append(f"Duplicate headers found: {sorted(duplicates)}.")

    return warnings


def _formula_warnings(columns: list[ColumnDef]) -> list[str]:
    """Generate warnings for columns whose data cells contain formulas."""

    warnings: list[str] = []
    for column in columns:
        if column.is_formula:
            warnings.append(f"Column '{column.name}' contains formulas.")
    return warnings


def _json_type_for_column(column_type: ColumnType) -> str:
    """Map EMA column types to JSON schema primitive types."""

    if column_type is ColumnType.NUMBER:
        return "number"
    if column_type is ColumnType.BOOL:
        return "boolean"
    return "string"


__all__ = [
    "as_extraction_contract",
    "propose_schema",
    "validate_live_schema",
]
