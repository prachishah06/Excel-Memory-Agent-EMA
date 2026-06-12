"""Safe Excel writer utilities for EMA."""

from __future__ import annotations

import os
import shutil
import tempfile
from datetime import UTC, datetime
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils.cell import get_column_letter, range_boundaries
from openpyxl.workbook.workbook import Workbook

from ema import config
from ema.errors import (
    EmaError,
    SheetNotFoundError,
    UndoError,
    ValidationError,
    WorkbookFileMissingError,
    WorkbookLockedError,
    WriteVerificationError,
)
from ema.models import AppendResult, ColumnDef, ColumnType, UndoResult, WorkbookSchema
from ema.schema import validate_live_schema


def check_writable(path: Path) -> None:
    """Validate that a workbook path exists, is unlocked, and opens in openpyxl.

    Raises:
        WorkbookFileMissingError: The path does not exist or is not a file.
        WorkbookLockedError: Excel or another process appears to have the file locked.
        EmaError: The file exists but is not a valid readable Excel workbook.
    """

    if not path.exists() or not path.is_file():
        raise WorkbookFileMissingError(f"Workbook file does not exist: {path}")

    lock_path = path.with_name(f"{config.LOCK_SUFFIX}{path.name}")
    if lock_path.exists():
        raise WorkbookLockedError(f"Workbook appears to be locked by Excel: {path}")

    try:
        with path.open("rb+"):
            pass
    except PermissionError as exc:
        raise WorkbookLockedError(f"Workbook is locked or not writable: {path}") from exc

    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise EmaError(f"Workbook is not a valid readable .xlsx file: {path}") from exc
    else:
        workbook.close()


def backup(path: Path) -> Path:
    """Create a byte-for-byte backup copy of a workbook and return its path.

    The backup is written beneath `config.BACKUP_DIR` in a workbook-specific
    subdirectory named after the workbook stem.

    Raises:
        WorkbookFileMissingError: The workbook path does not exist.
        WorkbookLockedError: The workbook is locked.
        EmaError: The workbook exists but is not a readable `.xlsx` file.
    """

    check_writable(path)

    workbook_backup_dir = config.BACKUP_DIR / path.stem
    workbook_backup_dir.mkdir(parents=True, exist_ok=True)

    backup_path = _next_backup_path(workbook_backup_dir, path.stem)
    shutil.copyfile(path, backup_path)
    return backup_path


def _next_backup_path(backup_dir: Path, workbook_stem: str) -> Path:
    """Return a unique timestamped backup path within a target directory."""

    while True:
        timestamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%S%fZ")
        candidate = backup_dir / f"{workbook_stem}-{timestamp}.xlsx"
        if not candidate.exists():
            return candidate


def append_row(path: Path, schema: WorkbookSchema, row: dict[str, object]) -> AppendResult:
    """Append one validated row to either a plain sheet or an Excel Table."""

    if schema.table_name is None:
        return _append_plain_row(path, schema, row)
    return _append_table_row(path, schema, row)


def verify_write(
    path: Path,
    schema: WorkbookSchema,
    expected: dict[str, object],
    row_idx: int,
) -> None:
    """Verify that a specific workbook row matches expected schema-keyed values.

    This check is read-only and ignores schema columns marked as formulas.
    Raises `WriteVerificationError` if the row is missing or any checked value
    differs from the workbook contents.
    """

    workbook = load_workbook(path, data_only=False)
    try:
        try:
            worksheet = workbook[schema.sheet]
        except KeyError as exc:
            raise SheetNotFoundError(f"Worksheet not found: {schema.sheet}") from exc

        if row_idx < 1 or row_idx > worksheet.max_row:
            raise WriteVerificationError(
                f"Row {row_idx} does not exist in worksheet '{schema.sheet}'."
            )

        columns_by_name = {column.name: column for column in schema.columns}

        for column_name, expected_value in expected.items():
            column = columns_by_name.get(column_name)
            if column is None:
                continue
            if column.is_formula:
                continue

            actual_value = worksheet.cell(row=row_idx, column=column.index + 1).value
            if actual_value != expected_value:
                raise WriteVerificationError(
                    f"Verification failed for row {row_idx}, column '{column_name}': "
                    f"expected {expected_value!r}, found {actual_value!r}."
                )
    finally:
        workbook.close()


def atomic_save(workbook: Workbook, path: Path) -> Path:
    """Save a workbook to a sibling temp file and replace the target atomically.

    The workbook is first saved to a temporary `.xlsx` file in the same
    directory, then re-opened to confirm it is readable before `os.replace()`
    swaps it into place. Any temporary file is removed on success or failure.
    """

    path.parent.mkdir(parents=True, exist_ok=True)

    file_descriptor, temp_name = tempfile.mkstemp(
        dir=path.parent,
        prefix=f"{path.stem}-tmp-",
        suffix=".xlsx",
    )
    os.close(file_descriptor)
    temp_path = Path(temp_name)

    try:
        workbook.save(temp_path)
        verified_workbook = load_workbook(temp_path, read_only=True, data_only=False)
        verified_workbook.close()
        os.replace(temp_path, path)
        return path
    except Exception:
        if temp_path.exists():
            temp_path.unlink()
        raise


def undo_last(path: Path, undo_token: str) -> UndoResult:
    """Restore a workbook from a specific backup path/token.

    The backup file is treated as the source of truth. Its bytes are copied to a
    sibling temp file, the temp workbook is verified to open successfully, and
    the target workbook is atomically replaced.
    """

    check_writable(path)
    backup_path = _resolve_backup_path(undo_token)

    file_descriptor, temp_name = tempfile.mkstemp(
        dir=path.parent,
        prefix=f"{path.stem}-undo-",
        suffix=".xlsx",
    )
    os.close(file_descriptor)
    temp_path = Path(temp_name)

    try:
        shutil.copyfile(backup_path, temp_path)
        verified_workbook = load_workbook(temp_path, read_only=True, data_only=False)
        verified_workbook.close()
        os.replace(temp_path, path)
    except Exception as exc:
        if temp_path.exists():
            temp_path.unlink()
        if isinstance(exc, UndoError):
            raise
        raise UndoError(f"Failed to restore workbook from backup: {backup_path}") from exc

    return UndoResult(
        ok=True,
        workbook_id=path.stem,
        restored_from=str(backup_path),
        message=f"Restored workbook from backup '{backup_path}'.",
    )


def table_append(path: Path, schema: WorkbookSchema, row: dict[str, object]) -> AppendResult:
    """Temporary compatibility wrapper for legacy direct table callers."""

    return _append_table_row(path, schema, row)


def _resolve_backup_path(undo_token: str) -> Path:
    """Resolve and validate an undo token/path."""

    candidate = Path(undo_token)
    if not undo_token.strip():
        raise UndoError("Undo token is empty.")

    if not candidate.is_absolute():
        candidate = config.BACKUP_DIR / candidate

    if not candidate.exists() or not candidate.is_file():
        raise UndoError(f"Backup file does not exist: {undo_token}")

    return candidate


def _append_plain_row(path: Path, schema: WorkbookSchema, row: dict[str, object]) -> AppendResult:
    """Append a validated row to a plain worksheet path."""

    backup_path, undo_token = _prepare_append(path, schema, row)

    workbook = load_workbook(path)
    try:
        worksheet = _get_worksheet(workbook, schema)
        target_row = _next_writable_row(worksheet, schema)
        ordered_preview = _write_row_values(worksheet, schema, row, target_row, 1)
        atomic_save(workbook, path)
    finally:
        workbook.close()

    return _finalize_append(
        path=path,
        schema=schema,
        target_row=target_row,
        ordered_preview=ordered_preview,
        backup_path=backup_path,
        undo_token=undo_token,
        message=f"Appended row {target_row} to '{schema.sheet}'.",
    )


def _append_table_row(path: Path, schema: WorkbookSchema, row: dict[str, object]) -> AppendResult:
    """Append a validated row to an Excel Table and expand its range."""

    backup_path, undo_token = _prepare_append(path, schema, row)

    if schema.table_name is None:
        raise ValidationError("Table append requires a schema with table_name set.")

    workbook = load_workbook(path)
    try:
        worksheet = _get_worksheet(workbook, schema)
        table = worksheet.tables[schema.table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        target_row = max_row + 1
        ordered_preview = _write_row_values(worksheet, schema, row, target_row, min_col)
        table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{target_row}"
        atomic_save(workbook, path)
    finally:
        workbook.close()

    return _finalize_append(
        path=path,
        schema=schema,
        target_row=target_row,
        ordered_preview=ordered_preview,
        backup_path=backup_path,
        undo_token=undo_token,
        message=f"Appended row {target_row} to table '{schema.table_name}'.",
    )


def _prepare_append(
    path: Path,
    schema: WorkbookSchema,
    row: dict[str, object],
) -> tuple[Path, str]:
    """Run the shared safety and validation checks required before any append."""

    check_writable(path)
    validate_live_schema(path, schema)
    _validate_row_input(schema, row)
    backup_path = backup(path)
    return backup_path, str(backup_path)


def _finalize_append(
    *,
    path: Path,
    schema: WorkbookSchema,
    target_row: int,
    ordered_preview: dict[str, object],
    backup_path: Path,
    undo_token: str,
    message: str,
) -> AppendResult:
    """Verify a completed write and build the public append result."""

    try:
        verify_write(path, schema, ordered_preview, target_row)
    except WriteVerificationError:
        undo_last(path, undo_token)
        raise

    return AppendResult(
        ok=True,
        workbook_id=path.stem,
        sheet=schema.sheet,
        written_row=target_row,
        row_preview=ordered_preview,
        backup_path=str(backup_path),
        undo_token=undo_token,
        dry_run=False,
        message=message,
    )


def _get_worksheet(workbook: Workbook, schema: WorkbookSchema):
    """Return a worksheet by persisted schema name or raise a typed error."""

    try:
        return workbook[schema.sheet]
    except KeyError as exc:
        raise SheetNotFoundError(f"Worksheet not found: {schema.sheet}") from exc


def _write_row_values(
    worksheet: object,
    schema: WorkbookSchema,
    row: dict[str, object],
    target_row: int,
    start_column: int,
) -> dict[str, object]:
    """Write schema-ordered row values starting at an absolute worksheet column."""

    ordered_preview: dict[str, object] = {}

    for column in schema.columns:
        if column.name not in row:
            continue

        value = _validated_cell_value(column, row[column.name])
        worksheet.cell(row=target_row, column=start_column + column.index, value=value)
        ordered_preview[column.name] = value

    return ordered_preview


def _validate_row_input(schema: WorkbookSchema, row: dict[str, object]) -> None:
    """Validate append input keys against the persisted workbook schema."""

    columns_by_name = {column.name: column for column in schema.columns}
    unknown_columns = set(row) - set(columns_by_name)
    if unknown_columns:
        raise ValidationError(
            f"Unknown column(s) for sheet '{schema.sheet}': {sorted(unknown_columns)}"
        )

    missing_required = [
        column.name
        for column in schema.columns
        if column.required and not column.is_formula and column.name not in row
    ]
    if missing_required:
        raise ValidationError(
            f"Missing required column(s) for sheet '{schema.sheet}': {missing_required}"
        )

    formula_columns = [
        column.name for column in schema.columns if column.is_formula and column.name in row
    ]
    if formula_columns:
        raise ValidationError(
            f"Cannot write to formula column(s): {formula_columns}"
        )


def _validated_cell_value(column: ColumnDef, value: object) -> object:
    """Validate a row value against the declared column type."""

    if value is None:
        return None

    if column.type is ColumnType.NUMBER:
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            raise ValidationError(f"Column '{column.name}' expects a numeric value.")
        return value

    if column.type is ColumnType.BOOL:
        if not isinstance(value, bool):
            raise ValidationError(f"Column '{column.name}' expects a boolean value.")
        return value

    if column.type is ColumnType.TEXT:
        if not isinstance(value, str):
            raise ValidationError(f"Column '{column.name}' expects a text value.")
        return value

    if column.type is ColumnType.DATE:
        if not isinstance(value, str):
            raise ValidationError(
                f"Column '{column.name}' expects an ISO date string for this step."
            )
        return value

    return value


def _next_writable_row(worksheet: object, schema: WorkbookSchema) -> int:
    """Compute the first empty row after the last populated data row."""

    last_populated_row = schema.data_start_row - 1

    for row_index in range(schema.data_start_row, worksheet.max_row + 1):
        if any(
            worksheet.cell(row=row_index, column=column.index + 1).value is not None
            for column in schema.columns
        ):
            last_populated_row = row_index

    return last_populated_row + 1
